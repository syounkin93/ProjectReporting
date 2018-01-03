##Built using Python 3.6.3
##Created by: Steven Younkin
##Last Modified 12/29/17

import glob, os, string
from openpyxl import Workbook, load_workbook
from openpyxl.styles import NamedStyle, Font, Border, Side, PatternFill, Color
from openpyxl.formatting.rule import ColorScale, FormatObject, ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.styles.differential import DifferentialStyle
from copy import copy
import fnmatch
import re
from time import gmtime, strftime
from InsertRow import insert_rows
import subprocess

path=os.path.dirname(os.path.realpath(__file__))
os.chdir(path)

##Set Global Variables
ProjectReportWorkbook = ""; projectreportfile = ""; FirstSheetProjectReport = ""
ProjectReportProjectManager = []; ProjectReportProjectNumber = []
OutX = int(); OutY = int()

##Find Current Quarter and Month:
CurrentMonth = int(strftime("%m", gmtime()))
CurrentQuarter = ""
if CurrentMonth > 9:
    CurrentQuarter = 4
if CurrentMonth > 6 and CurrentMonth < 10:
    CurrentQuarter = 3
if CurrentMonth > 3 and CurrentMonth < 7:
    CurrentQuarter = 2
if CurrentMonth < 4:
    CurrentQuarter = 1

##If End of year, uncomment next line. This will let all quarters run
##CurrentQuarter = 5
    
#Step 1: Read values from Project Report
for projectreport in glob.glob("Water Team Project Report - New Format*"):
    projectreportfile = projectreport
    ProjectReportWorkbook = load_workbook(projectreportfile)
    FirstSheetProjectReport = ProjectReportWorkbook["WATER Report"]

    FirstSheetProjectReport.conditional_formatting.cf_rules == ""
    
    for column in FirstSheetProjectReport[1]:
        try:
            if re.search("Project Manager",column.value):
                PMcolumn = column.column
            if re.search("Project Description",column.value):
                ProjectDescriptionColumn = column.column
            if re.search("Project Number",column.value):
                ProjectNumberColumn = column.column
            if re.search("Net Total Contract",column.value):
                NetTotalContractColumn = column.column
            if re.search("Gross Billed",column.value):
                GrossBilledColumn = column.column
            if re.search("Total Revenue",column.value):
                TotalRevenue = column.column
        except TypeError:
            break
            
    #Find Columns for months in this year in Project Report      
    for column in FirstSheetProjectReport['A1':TotalRevenue][0]:
        if re.search("Net Revenue - January",column.value) is not None:
            ThisJanuaryColumn = column.column
        if re.search("Net Revenue - February",column.value) is not None:
            ThisFebruaryColumn = column.column
        if re.search("Net Revenue - March",column.value) is not None:
            ThisMarchColumn = column.column
        if re.search("Net Revenue - April",column.value) is not None:
            ThisAprilColumn = column.column
        if re.search("Net Revenue - May",column.value) is not None:
            ThisMayColumn = column.column
        if re.search("Net Revenue - June",column.value) is not None:
            ThisJuneColumn = column.column
        if re.search("Net Revenue - July",column.value) is not None:
            ThisJulyColumn = column.column
        if re.search("Net Revenue - August",column.value) is not None:
            ThisAugustColumn = column.column
        if re.search("Net Revenue - September",column.value) is not None:
            ThisSeptemberColumn = column.column
        if re.search("Net Revenue - October",column.value) is not None:
            ThisOctoberColumn = column.column
        if re.search("Net Revenue - November",column.value) is not None:
            ThisNovemberColumn = column.column
        if re.search("Net Revenue - December",column.value) is not None:
            ThisDecemberColumn = column.column
        if re.search("Q1 Net Revenue",column.value) is not None:
            Q1Net = column.column
        if re.search("Q2 Net Revenue",column.value) is not None:
            Q2Net = column.column
        if re.search("Q3 Net Revenue",column.value) is not None:
            Q3Net = column.column
        if re.search("Q4 Net Revenue",column.value) is not None:
            Q4Net = column.column
        if re.search("Q1 Net Projected",column.value) is not None:
            Q1Projected = column.column
        if re.search("Q2 Net Projected",column.value) is not None:
            Q2Projected = column.column
        if re.search("Q3 Net Projected",column.value) is not None:
            Q3Projected = column.column
        if re.search("Q4 Net Projected",column.value) is not None:
            Q4Projected = column.column

    #Create List of Quarterly Revenue
    QuarterlyRevenue = [Q1Net,Q2Net,Q3Net,Q4Net]
                
    #Create List of Quarterly Projected
    QuarterlyProjectionColumns = [Q1Projected,Q2Projected,Q3Projected,Q4Projected]

    #Create List of Quarters to current completed Quarter
    CurrentQuarterlyProjectionColumns = QuarterlyProjectionColumns[:CurrentQuarter-1]
    CurrentQuarterlyRevenue = QuarterlyRevenue[:CurrentQuarter-1]
    
    ###Create QA/QC Sheet
    try:
        ProjectReportWorkbook["QAQC Checks"]
        QAQCsheet = ProjectReportWorkbook["QAQC Checks"]
    except KeyError:
        ProjectReportWorkbook.create_sheet("QAQC Checks")
        QAQCsheet = ProjectReportWorkbook["QAQC Checks"]
        QAQCsheet["F1"].value = 0
    QAQCsheet['A1'] = "Unmatched Revenue Sheet Projects"
    for x in QAQCsheet['A1:L1'][0]:
        x.fill = PatternFill("solid", fgColor="ffff00")
        
    #Add Run Times
    Runtimes = QAQCsheet["F1"].value
    QAQCrow = 2
    QAQCcol = Runtimes*10 + 1

##Step 2: Find Projected Values. This script finds values from 2017, which only October to December are Projected.
Projections = []; ProjectionsSheet=""
    
for file in glob.glob("Water Team - Project Report*.xlsx"):
    print("Projections: " + file)
    projectionsreport = file
    ProjectionsWorkbook = load_workbook(projectionsreport, data_only=True)
    ProjectionsSheet = ProjectionsWorkbook["2017 Projections"]
    Total2017col="";Total2018col=""
    for column in ProjectionsSheet['A9:' +  ProjectionsSheet.cell(row=9,column=ProjectionsSheet.max_column).coordinate][0]:
        if column.value == "Total 2016":
            Total2016col = column.column
        if column.value == "Total 2017":
            Total2017col = column.column
        if column.value == "Total 2018":
            Total2018col = column.column  
        
    for row in ProjectionsSheet["A11:A"+str(ProjectionsSheet.max_row-80)]:
        row = row[0]
        try:
            if row.font.b == True:
                ProjectManager = row.value
                pass
        except TypeError:
            pass
        
        if row.offset(row=0,column=1).value is not None:
            ProjectNumber = row.offset(row=0,column=1).value
            ProjectName = row.value
            ProjRow = Total2016col+str(row.row)
            ProjectJan = ProjectionsSheet[ProjRow].offset(row=0,column=1).value
            ProjectFeb = ProjectionsSheet[ProjRow].offset(row=0,column=2).value
            ProjectMar = ProjectionsSheet[ProjRow].offset(row=0,column=3).value
            ProjectApr = ProjectionsSheet[ProjRow].offset(row=0,column=5).value
            ProjectMay = ProjectionsSheet[ProjRow].offset(row=0,column=6).value
            ProjectJune = ProjectionsSheet[ProjRow].offset(row=0,column=7).value
            ProjectJuly = ProjectionsSheet[ProjRow].offset(row=0,column=9).value
            ProjectAug = ProjectionsSheet[ProjRow].offset(row=0,column=10).value
            ProjectSep = ProjectionsSheet[ProjRow].offset(row=0,column=11).value
            ProjectOct = ProjectionsSheet[ProjRow].offset(row=0,column=13).value
            ProjectNov = ProjectionsSheet[ProjRow].offset(row=0,column=14).value
            ProjectDec = ProjectionsSheet[ProjRow].offset(row=0,column=15).value
            
            Projections.append([ProjectManager,ProjectNumber,ProjectName,ProjectJan
                                ,ProjectFeb,ProjectMar,ProjectApr,ProjectMay,ProjectJune,
                                ProjectJuly,ProjectAug,ProjectSep,ProjectOct,ProjectNov,
                                ProjectDec])
            
            
            
        if row.offset(row=0,column=1).value is None and row.font.b == False:
            print(str(row.offset(row=0,column=0).value)  + " Missing PN: " + str(row.offset(row=0,column=1).value))
        
        
##Step 3: Iterate Through Principal Summary Reports from current working directory
for file in glob.glob("*Active Projects_*"):
    print("Revenue from Active Projects: " + file)
    RevenueFilePath = os.path.join(path,file)
            
    RevenueWorkbook = load_workbook(RevenueFilePath)
    ActiveSheet = RevenueWorkbook["Principal Summary Report- WATER"]
    #Find Principal Summary Report Columns
    TotGrossCol = ""; BilledCol = "";ReceivedCol ="";AROutstandingCol ="";RemainingCol ="";ExpDirectReimburseCol ="";NetRevenueCol=""
    for col in ActiveSheet['9']:
        if col.value is not None:
            if bool(re.search("Total \(Gross",col.value)):
                TotGrossCol = col.column
            if bool(re.search("Billed",col.value)):
                BilledCol = col.column
            if bool(re.search("Received",col.value)):
                ReceivedCol = col.column
            if bool(re.search("A/R",col.value)):
                AROutstandingCol = col.column
            if bool(re.search("Remaining",col.value)):
                RemainingCol = col.column
            if bool(re.search("Exp",col.value)):
                ExpDirectReimburseCol = col.column
            if bool(re.search("Revenue",col.value)):
                NetRevenueCol = col.column
                           
    
    #Find Month of Report and set output Column to that Value                
    monthString = ActiveSheet.cell(row=6,column=6).value
    thirdSpace = [m.start() for m in re.finditer(" ",monthString)][2]
    firstSlash = [m.start() for m in re.finditer("/",monthString)][0]
    month = monthString[thirdSpace+1:firstSlash]

    if month == "1":
        OutX = ThisJanuaryColumn
    if month == "2":
        OutX = ThisFebruaryColumn
    if month == "3":
        OutX = ThisMarchColumn
    if month == "4":
        OutX = ThisAprilColumn
    if month == "5":
        OutX = ThisMayColumn
    if month == "6":
        OutX = ThisJuneColumn
    if month == "7":
        OutX = ThisJulyColumn
    if month == "8":
        OutX = ThisAugustColumn
    if month == "9":
        OutX = ThisSeptemberColumn
    if month == "10":
        OutX = ThisOctoberColumn
    if month == "11":
        OutX = ThisNovemberColumn
    if month == "12":
        OutX = ThisDecemberColumn

    RevenueSheet = []
    
    #Build RevenueSheet, which is list of lists representing projects identified
    #by PM and project number
    for row in ActiveSheet['A']:
        try:
            if re.search("PM Name:*",row.value) is not None:
                ProjectManager = row.value[9:]
                PMFirstName = ProjectManager[ProjectManager.index(",")+2:]
                PMLastName = ProjectManager[:ProjectManager.index(",")]
                PMFirstLastName = PMFirstName + " " + PMLastName #0
            if re.search("\d{8,9}",row.value) is not None:
                ##print row.value
                ProjectNumber = row.value[0:12].lstrip("0") #1
                ProjectName = row.value[13:] #2
                ProjectContract = row.coordinate(TotGrossCol+str(row.row)).value #3
                ProjectBilled = row.coordinate(BilledCol+str(row.row)).value #4
                ProjectReceived = row.coordinate(ReceivedCol+str(row.row)).value #5
                ProjectRemaining = row.coordinate(RemainingCol+str(row.row)).value #6
                ProjectAROutstanding = row.coordinate(AROutstandingCol+str(row.row)).value #7
                ProjectConsultantReimbursement = row.coordinate(ExpDirectReimburseCol+str(row.row)).value #8
                ProjectNetRevenue = row.coordinate(TotGrossCol+str(row.row)).value #9
                RevenueSheet.append([PMFirstLastName,ProjectNumber,ProjectName,ProjectContract,ProjectBilled,ProjectReceived,ProjectRemaining,ProjectAROutstanding,ProjectConsultantReimbursement,ProjectNetRevenue])
        except TypeError:
            pass
        
    #ConnectRevenueSheet to Project Report Months
    QAQCrow += 1
    QAQCtime = strftime("%Y-%m-%d %H:%M:%S", gmtime())
    
    QAQCsheet["H1"].value = QAQCtime = "Revenue Month: " + month + " Run time: " + strftime("%Y-%m-%d %H:%M:%S", gmtime())
    QAQCsheet.cell(row=QAQCrow,column=QAQCcol).value = "Month: " + month + " | " + str(QAQCtime)
    QAQCrow += 2
    
    for RevenueRow in RevenueSheet:
        matched = "Unmatched"
        for ProjectReportRow in range(2,FirstSheetProjectReport.max_row):
            PMcoord = PMcolumn + str(ProjectReportRow)
            PNcoord = ProjectNumberColumn + str(ProjectReportRow)
            Monthcoord = OutX + str(ProjectReportRow)
            if RevenueRow[0] == FirstSheetProjectReport[PMcoord].value and RevenueRow[1] == FirstSheetProjectReport[PNcoord].value:
                matched = "Match"
                FirstSheetProjectReport[Monthcoord].value = RevenueRow[8]

                #Filling Generic Project Values
                FirstSheetProjectReport[Q1Net+str(ProjectReportRow)].value = "=sum(" + ThisJanuaryColumn + str(ProjectReportRow) + ":" + ThisMarchColumn + str(ProjectReportRow)+")"
                FirstSheetProjectReport[Q2Net+str(ProjectReportRow)].value = "=sum(" + ThisAprilColumn + str(ProjectReportRow) + ":" + ThisJuneColumn + str(ProjectReportRow)+")"
                FirstSheetProjectReport[Q3Net+str(ProjectReportRow)].value = "=sum(" + ThisJulyColumn + str(ProjectReportRow) + ":" + ThisSeptemberColumn + str(ProjectReportRow)+")"
                FirstSheetProjectReport[Q4Net+str(ProjectReportRow)].value = "=sum(" + ThisOctoberColumn + str(ProjectReportRow) + ":" + ThisDecemberColumn + str(ProjectReportRow)+")"
   
            else:
                pass
        if matched == "Unmatched":
            #insert after last instance of PM
            try:
                LastPM = max(loc for loc, val in enumerate(FirstSheetProjectReport[PMcolumn]) if val.value == RevenueRow[0])
            except ValueError:
                LastPM = FirstSheetProjectReport.max_row

            FirstSheetProjectReport.insert_rows(row_idx=LastPM,cnt=1,copy_style=True,fill_formulae=True)
            FirstSheetProjectReport[ProjectNumberColumn + str(LastPM+1)].value = RevenueRow[1]
            FirstSheetProjectReport[PMcolumn + str(LastPM+1)].value = RevenueRow[0]
            FirstSheetProjectReport[ProjectDescriptionColumn +str(LastPM+1)].value = RevenueRow[2]
            FirstSheetProjectReport[OutX + str(LastPM+1)].value = RevenueRow[9]

            FirstSheetProjectReport[Q1Net+str(LastPM+1)].value = "=sum(" + ThisJanuaryColumn + str(LastPM+1) + ":" + ThisMarchColumn + str(LastPM+1)+")"
            FirstSheetProjectReport[Q2Net+str(LastPM+1)].value = "=sum(" + ThisAprilColumn + str(LastPM+1) + ":" + ThisJuneColumn + str(LastPM+1)+")"
            FirstSheetProjectReport[Q3Net+str(LastPM+1)].value = "=sum(" + ThisJulyColumn + str(LastPM+1) + ":" + ThisSeptemberColumn + str(LastPM+1)+")"
            FirstSheetProjectReport[Q4Net+str(LastPM+1)].value = "=sum(" + ThisOctoberColumn + str(LastPM+1) + ":" + ThisDecemberColumn + str(LastPM+1)+")"

                    
            QAQCsheet.cell(row=QAQCrow,column=QAQCcol).value = "Revenue Unmatched: " + RevenueRow[0] + " " + RevenueRow[1] + " " + RevenueRow[2]
            QAQCrow += 1

#Adding New row for Project Managers
lt = Side(style='thin',color="000000")
bd = Side(style='thick',color="000000")
PMlist = []
	
for PM in range(2,len(FirstSheetProjectReport['A'])+12):
    currentPM = FirstSheetProjectReport['A'+str(PM)].value
    pastPM = FirstSheetProjectReport['A'+str(PM-1)].value
    if currentPM is not pastPM and pastPM is not None and currentPM is not None and currentPM not in PMlist:
        try:
            print(str(pastPM) + " " + str(currentPM) + " " + str(PM))
            FirstSheetProjectReport.insert_rows(row_idx=PM,cnt=1,above=True,copy_style=False,fill_formulae=False)
            insertedPM = FirstSheetProjectReport["B" + str(PM)]
            insertedPM.value = currentPM
            insertedPM.font = Font(b=True,size=14)
            insertedPM.border = Border(left=lt,top=lt,right=lt,bottom=bd)
            insertedPM.fill = PatternFill("solid",fgColor="BFBFBF")
            FirstSheetProjectReport.row_dimensions[PM].height = 18
            FirstSheetProjectReport.merge_cells(start_row=PM,start_column=2,end_row=PM,end_column=28)
            
            #Add This PM to list to ensure it is not duplicated
            PMlist.append(currentPM)
             
        except TypeError:
            print(str(PM) + " Typeerror")
            pass
    else:
        pass
    
#Add Quarterly Projections
ProjectedPM = "ProjectedPM"

redFill = PatternFill(start_color='EE1111',
                end_color='EE1111',
                fill_type='solid')

greenFill = PatternFill(start_color='11ee11',
                end_color='11ee11',
                fill_type='solid')

blueFill = PatternFill(start_color='1111ee',
                end_color='1111ee',
                fill_type='solid')

for ProjectReportRow in range(2,FirstSheetProjectReport.max_row):
    ProjReportRow = str(ProjectReportRow)
    if FirstSheetProjectReport["B"+ProjReportRow].font.b == True:
        ProjectedPM = FirstSheetProjectReport["B"+ProjReportRow].value
                
    for ProjectedRow in Projections:
        ProjectedRow = [0 if x is None else x for x in ProjectedRow]

        if ProjectedRow[0] ==  ProjectedPM and ProjectedRow[1] == FirstSheetProjectReport["C"+ProjReportRow].value:
            
            i = 3
            Q = 1
            for Quar in QuarterlyProjectionColumns:

                QuarCoord = FirstSheetProjectReport[Quar+ProjReportRow].offset(row=0,column=-1).coordinate
                QuarProjectionCoord = FirstSheetProjectReport[Quar+ProjReportRow].coordinate

                FirstSheetProjectReport[Quar+ProjReportRow].value = ProjectedRow[i] + ProjectedRow[i+1] + ProjectedRow[i+2]

                if Q < CurrentQuarter:
                    FirstSheetProjectReport.conditional_formatting.add(QuarCoord,
                    CellIsRule(operator='lessThan',formula=[QuarProjectionCoord],stopIfTrue=True,fill = redFill))
                    FirstSheetProjectReport.conditional_formatting.add(QuarCoord,
                    CellIsRule(operator='equal',formula=[QuarProjectionCoord],stopIfTrue=True,fill = blueFill))
                    FirstSheetProjectReport.conditional_formatting.add(QuarCoord,
                    CellIsRule(operator='greaterThan',formula=[QuarProjectionCoord],stopIfTrue=True,fill = greenFill))
                    
                i += 3
                Q += 1
       

        
print("Saving to:" + path + "\\" + projectreportfile)
QAQCsheet["F1"].value = QAQCsheet["F1"].value + 1
QAQCsheet.sheet_view.zoomScale = 55
ProjectReportWorkbook.save(projectreportfile)
os.startfile(path + "\\" + projectreportfile)
